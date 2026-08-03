import csv
import io
import unicodedata
import xml.etree.ElementTree as ET
from datetime import date, datetime

from django.db import transaction
from openpyxl import Workbook, load_workbook

from .models import Client


COLUMNS = [
    ("name", "Nome"), ("document", "CNPJ_CPF"), ("email", "Email"), ("phone", "Telefone"),
    ("process_number", "Numero_Processo"), ("publication_date", "Data_Publicacao"),
    ("street", "Rua"), ("address_number", "Numero"), ("complement", "Complemento"),
    ("neighborhood", "Bairro"), ("city", "Cidade"), ("state", "Estado"), ("postal_code", "CEP"),
    ("notification_number", "Numero_Notificacao"), ("classification", "Classificacao"),
    ("action_description", "Descricao_Providencia"), ("validation", "Validacao"), ("active", "Ativo"),
]


def _key(value):
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode().lower()
    return "".join(char for char in text if char.isalnum())


ALIASES = {_key(label): field for field, label in COLUMNS}
ALIASES.update({_key(field): field for field, _ in COLUMNS})
ALIASES.update({
    "cnpjcpf": "document", "condominio": "name", "nomeidentificacao": "name",
    "processo": "process_number", "numerodoprocesso": "process_number",
    "datadapublicacao": "publication_date", "notificacao": "notification_number",
    "numerodanotificacao": "notification_number", "endereco": "street",
    "descricaoprovidencia": "action_description",
})


def _date(value):
    if not value:
        return None
    if isinstance(value, (date, datetime)):
        return value.date() if isinstance(value, datetime) else value
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Data inválida: {value}")


def _bool(value):
    return _key(value) not in {"0", "nao", "false", "inativo"}


def _normalize(row):
    data = {ALIASES.get(_key(key)): value for key, value in row.items() if ALIASES.get(_key(key))}
    data = {key: (str(value).strip() if value is not None else "") for key, value in data.items()}
    data["publication_date"] = _date(data.get("publication_date"))
    data["active"] = _bool(data.get("active", "sim"))
    data["state"] = data.get("state", "RJ").upper()[:2]
    if not data.get("name"):
        raise ValueError("Nome do condomínio não informado")
    combined_address = data.get("street", "")
    if combined_address and not data.get("address_number"):
        import re
        match = re.match(r"^(.*?)(?:,|\s+n[º°o.]*)\s*(\d+[A-Za-z]?)\s*$", combined_address, re.IGNORECASE)
        if match:
            data["street"], data["address_number"] = match.group(1).strip(), match.group(2)
    classification_key = _key(data.get("classification"))
    if "autovistoria" in classification_key:
        data["classification"] = "autovistoria"
    elif "manutencao" in classification_key:
        data["classification"] = "manutencao_predial"
    elif data.get("classification") not in dict(Client.CLASSIFICATIONS):
        labels = {_key(label): value for value, label in Client.CLASSIFICATIONS}
        data["classification"] = labels.get(classification_key, "outros")
    if data.get("validation") not in dict(Client.VALIDATIONS):
        labels = {_key(label): value for value, label in Client.VALIDATIONS}
        data["validation"] = labels.get(_key(data.get("validation")), "validar")
    return data


def read_clients(upload):
    extension = upload.name.rsplit(".", 1)[-1].lower()
    if extension == "xlsx":
        sheet = load_workbook(upload, read_only=True, data_only=True).active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "") for value in next(rows)]
        return [dict(zip(headers, values)) for values in rows if any(value not in (None, "") for value in values)]
    raw = upload.read()
    if extension in {"csv", "txt"}:
        text = raw.decode("utf-8-sig")
        sample = text[:4096]
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=";,\t|").delimiter
        except csv.Error:
            delimiter = ";"
        return list(csv.DictReader(io.StringIO(text), delimiter=delimiter))
    if extension == "xml":
        root = ET.fromstring(raw)
        return [{child.tag: child.text or "" for child in record} for record in root.findall(".//registro")]
    raise ValueError("Formato não permitido. Use XLSX, CSV, TXT ou XML.")


@transaction.atomic
def import_clients(upload):
    created = updated = 0
    errors = []
    for number, row in enumerate(read_clients(upload), start=2):
        try:
            data = _normalize(row)
            notification = data.get("notification_number", "")
            process = data.get("process_number", "")
            match = Client.objects.filter(notification_number__iexact=notification, process_number__iexact=process).first() if notification and process else None
            document = "".join(char for char in data.get("document", "") if char.isdigit())
            if not match and document:
                match = next((client for client in Client.objects.exclude(document="") if "".join(char for char in client.document if char.isdigit()) == document), None)
            if not match and data.get("street"):
                match = Client.objects.filter(
                    name__iexact=data["name"], street__iexact=data["street"],
                    address_number__iexact=data.get("address_number", ""),
                ).first()
            if match:
                for field, value in data.items():
                    setattr(match, field, value)
                match.save()
                updated += 1
            else:
                Client.objects.create(**data)
                created += 1
        except Exception as exc:
            errors.append(f"Linha {number}: {exc}")
    return created, updated, errors


def _rows():
    yield [label for _, label in COLUMNS]
    for client in Client.objects.all().order_by("name"):
        values = []
        for field, _ in COLUMNS:
            value = getattr(client, field)
            if isinstance(value, (date, datetime)):
                value = value.strftime("%d/%m/%Y")
            elif isinstance(value, bool):
                value = "Sim" if value else "Não"
            values.append(value or "")
        yield values


def export_clients(fmt):
    rows = list(_rows())
    if fmt == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Condomínios"
        for row in rows:
            sheet.append(row)
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = cell.font.copy(bold=True, color="FFFFFF")
            cell.fill = cell.fill.copy(fill_type="solid", fgColor="7427E8")
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if fmt in {"csv", "txt"}:
        output = io.StringIO()
        writer = csv.writer(output, delimiter=";" if fmt == "csv" else "\t", lineterminator="\n")
        writer.writerows(rows)
        return output.getvalue().encode("utf-8-sig"), "text/csv" if fmt == "csv" else "text/plain"
    if fmt == "xml":
        root = ET.Element("condominios")
        for values in rows[1:]:
            record = ET.SubElement(root, "registro")
            for (_, label), value in zip(COLUMNS, values):
                ET.SubElement(record, label).text = str(value)
        return ET.tostring(root, encoding="utf-8", xml_declaration=True), "application/xml"
    raise ValueError("Formato de exportação inválido.")
